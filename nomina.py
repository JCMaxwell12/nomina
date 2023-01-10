from openpyxl import load_workbook
from datetime import datetime
from datetime import date
from datetime import timedelta
import logging
import os

# crear nomina.log

logging.basicConfig(filename="nomina.log",
                    level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    filemode='w')

logger = logging.getLogger()

# verificar si existen y crear carpetas
if os.path.exists('input') is False:
    if os.path.isfile('input') is False:
        os.mkdir('input')
        logger.info('created input/')
if os.path.exists('output') is False:
    if os.path.isfile('output') is False:
        os.mkdir('output')
        logger.info('created output/')
wb = None
# Cargar el primer archivo en input/
for i in os.listdir('input'):
    if i.find('.xlsx') != -1:
        wb = load_workbook(os.path.join('input', i))
        break
if wb is None:
    logger.critical('Ningún archivo .xlsx encontrado en input/')

empleados = dict()
ids = list()
sheet = wb['Invoice']

# obtener dimensiones de hoja
num_empleados = sheet.calculate_dimension()
# cantidad de instancias (de checkin o checkout)
num_empleados = int(num_empleados[(2+num_empleados.find(':')):]) - 5
# loggear en caso de falla


class Empleado:
    """Empleado con número, ID, nombre, instancias de entrada o salida,
    cantidad de retardos y salidas anticipadas"""

    def __init__(self, num, id_nom, nombre):
        self.num, self.id_nom, self.nombre = num, id_nom, nombre.title()
        self.instancias = list()    # instancias de entrada o salida
        self.retardos = list()      # retardos
        self.anticipadas = list()   # salidas anticipadas
        self.trabajado = list()     # tiempo trabajado

    def __str__(self):
        x = (f'Empleado:\n   num: {self.num}\n   id: {self.id_nom}\n'
             f'   nombre: {self.nombre}\n'
             f'   retardos: {self.retardos}\n'
             f'   salidas anticipadas: {self.anticipadas}\n'
             f'   trabajado: {self.trabajado}\n'
             f'   tiempo: {self.tiempo()}\n'
             f'   instancias:\n')

        for i in range(len(self.instancias)):
            x += '    ' + str(self.instancias[i]) + '\n'
        return x

    def tiempo(self):
        self.val = timedelta(0)
        for i in self.trabajado:
            try:
#                if i < timedelta(0):
#                    i += timedelta(days=1)
                self.val += i
#                if i < timedelta(0):
#                    i += timedelta(days=1)

            except TypeError:
                self.val += timedelta(i)
        return self.val


# crear lista de empleados
for empleado in range(6, num_empleados + 5):
    # obtener valores
    emp_num = int(sheet['A'+str(empleado)].value)    # número de empleado
    emp_id = sheet['B'+str(empleado)].value     # id de nómina de empleado
    emp_nom = sheet['C'+str(empleado)].value    # nombre de empleado

    # agregar objeto de empleado si no está
    if sheet['A'+str(empleado)].value not in ids:
        empleados.update({emp_num: Empleado(emp_num, emp_id, emp_nom)})
        ids.append(sheet['A'+str(empleado)].value)

# añadir checkins y checkouts
    for i in empleados:
        entr = sheet['H'+str(empleado)].value
        sali = sheet['I'+str(empleado)].value
        chin = sheet['J'+str(empleado)].value
        chou = sheet['K'+str(empleado)].value
        fech = sheet['F'+str(empleado)].value

        if int(emp_num) == i:
            try:
                empleados[int(emp_num)].instancias.append({
                    'fecha':    datetime.strptime(fech, '%Y/%m/%d').date(),
                    'entrada':  datetime.strptime(entr+fech, '%I:%M %p%Y/%m/%d'),
                    'salida':   datetime.strptime(sali+fech, '%I:%M %p%Y/%m/%d'),
                    'checkin':  datetime.strptime(chin+fech, '%I:%M %p%Y/%m/%d'),
                    'checkout': datetime.strptime(chou+fech, '%I:%M %p%Y/%m/%d')})
            except TypeError:
                empleados[int(emp_num)].instancias.append({
                    'fecha':    datetime.strptime(fech, '%Y/%m/%d').date(),
                    'entrada':  datetime.strptime(entr+fech, '%I:%M %p%Y/%m/%d'),
                    'salida':   datetime.strptime(sali+fech, '%I:%M %p%Y/%m/%d'),
                    'checkin':  None,
                    'checkout': None})
                logger.warning(f'TypeError {fech, entr, sali, chin, chou}')
            # Corregir checkout para turno nocturno
            if empleados[int(emp_num)].instancias[-1]['entrada'] > empleados[int(emp_num)].instancias[-1]['salida']:
                empleados[int(emp_num)].instancias[-1]['salida'] += timedelta(days=1)
                try:
                    empleados[int(emp_num)].instancias[-1]['checkout'] += timedelta(days=1)
                except TypeError:
                    pass

# obtener horas trabajadas, inasistencias, retardos...
for empleado in empleados:
    for instancia in empleados[empleado].instancias:
        entr = instancia['entrada']
        sali = instancia['salida']
        try:
            chin = instancia['checkin']
        except AttributeError:
            chin = None
            # Si no hay checkin para la instancia
            logger.warning(f'Sin checkin ({instancia["checkin"]})')
        try:
            chou = instancia['checkout']
        except AttributeError:
            # Si no hay checkout para la instancia
            chou = None
            logger.warning(f'Sin checkout ({instancia["checkout"]})')

        if chou is None:  # tiempo trabajado
            logger.error(f'AttributeError: Error calculando tiempo trabajado'
                         f'\n   Checkout:    {instancia["checkout"]} = {chou}')

        elif chin is None:
            logger.error(f'AttributeError: Error calculando tiempo trabajado'
                         f'\n   Checkin:     {instancia["checkin"]} = {chin}')

        else:
            # tiempo trabajado
            empleados[empleado].trabajado.append(chou - chin)
            # agregar retardo
            if entr < chin:
                empleados[empleado].retardos.append(chin)

            # agregar salida anticipada
            if sali > chou:
                empleados[empleado].anticipadas.append(chou)
# Ordenar datos
empleados = {key: val for key, val in sorted(empleados.items(), key=lambda ele: ele[0])}

# Escribir datos
for empleado in empleados:
    # crear una hoja por empleado
    titulo = str(empleados[empleado].num)
    wb.create_sheet(title=titulo)
    sheet = wb[titulo]
    insts = empleados[empleado].instancias

    sheet['A1'], sheet['B1'] = 'Nombre', empleados[empleado].nombre
    sheet['C1'], sheet['D1'] = 'ID', empleados[empleado].id_nom
    sheet['E1'], sheet['F1'] = 'Num', empleados[empleado].num

    sheet['A3'] = 'Fecha'
    sheet['B3'] = 'Entrada'
    sheet['C3'] = 'Checkin'
    sheet['D3'] = 'Salida'
    sheet['E3'] = 'Checkout'
    sheet['F3'] = 'Horas trabajadas'

    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 18

    for i, instancia in enumerate(insts, 4):
        i = str(i)
        sheet['A'+i] = instancia['fecha']
        sheet['B'+i] = instancia['entrada']
        sheet['C'+i] = instancia['checkin']
        sheet['D'+i] = instancia['salida']
        sheet['E'+i] = instancia['checkout']
        try:
            sheet['F'+i] = instancia['checkout'] - instancia['checkin']
        except TypeError:
            sheet['F'+i] = None


# añadir a la hoja general
wb.create_sheet('Reporte', 1)
sheet = wb['Reporte']
sheet['A1'] = 'Número'
sheet['B1'] = 'Nombre'
sheet['C1'] = 'Retardos'
sheet['D1'] = 'Salidas anticipadas'
sheet['E1'] = 'Tiempo trabajado'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 40
sheet.column_dimensions['C'].width = 10
sheet.column_dimensions['D'].width = 10
sheet.column_dimensions['E'].width = 10

for i, empleado in enumerate(empleados, 2):
    i = str(i)
    sheet['A'+i] = empleados[empleado].num
    sheet['B'+i] = empleados[empleado].nombre
    sheet['C'+i] = len(empleados[empleado].retardos)
    sheet['D'+i] = len(empleados[empleado].anticipadas)
    sheet['E'+i] = empleados[empleado].tiempo()

savepath = str(date.today()) + ' Checkins and Checkouts.xlsx'
wb.save(os.path.join('output', savepath))
